"""Leak: read PV names from target PV sheet → write to target Leak sheet."""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.workbook.workbook import Workbook

from . import target_layout as L
from .columns import col_letter_to_index
from .excel_io import find_sheet
from .pressure_vessel import TransferMode, TransferReport
from .sheet_utils import clear_range, find_last_populated_row, iter_populated_rows


@dataclass
class LeakSizeEntry:
    """One leak size: a name and an optional orifice diameter."""

    name: str
    orifice_diameter: float | None  # None = not set (or FBR with source lookup)


@dataclass
class LeakSourceConfig:
    """Identifies where to find FBR max line sizes in the source workbook."""

    sheet: str
    name_col: str
    start_row: int
    max_line_size_col: str


@dataclass
class LeakOptions:
    leak_sizes: list[LeakSizeEntry]
    fbr_enabled: bool = False
    transfer_mode: TransferMode = TransferMode.OVERWRITE


@dataclass
class PVLeakRecord:
    """One pressure vessel row read from the target Pressure vessel sheet."""

    use: str
    study: str | None
    folder: str | None
    pv_name: str


# ---------------------------------------------------------------------------


def read_pv_names_from_target(
    tgt_wb: Workbook,
    report: TransferReport,
) -> list[PVLeakRecord]:
    """Read pressure vessel names (and Use/Study/Folder) from the target PV sheet."""
    ws = find_sheet(tgt_wb, L.PV_SHEET_NAME)
    use_idx = col_letter_to_index(L.PV_COL_USE)
    study_idx = col_letter_to_index(L.PV_COL_STUDY)
    folder_idx = col_letter_to_index(L.PV_COL_FOLDER)
    name_idx = col_letter_to_index(L.PV_COL_NAME)

    records: list[PVLeakRecord] = []
    for r in iter_populated_rows(ws, name_idx, L.DATA_START_ROW):
        use = ws.cell(row=r, column=use_idx).value
        study = ws.cell(row=r, column=study_idx).value
        folder = ws.cell(row=r, column=folder_idx).value
        name = ws.cell(row=r, column=name_idx).value
        records.append(PVLeakRecord(
            use=str(use).strip() if use is not None else L.LEAK_VALUE_USE,
            study=str(study).strip() if study is not None else None,
            folder=str(folder).strip() if folder is not None else None,
            pv_name=str(name).strip(),
        ))
    return records


def read_fbr_diameters(
    src_wb: Workbook,
    cfg: LeakSourceConfig,
    report: TransferReport,
) -> dict[str, float]:
    """Read max line sizes from the source workbook, keyed by PV name."""
    ws = find_sheet(src_wb, cfg.sheet)
    name_idx = col_letter_to_index(cfg.name_col)
    size_idx = col_letter_to_index(cfg.max_line_size_col)

    result: dict[str, float] = {}
    for r in iter_populated_rows(ws, name_idx, cfg.start_row):
        name = ws.cell(row=r, column=name_idx).value
        size = ws.cell(row=r, column=size_idx).value
        if name is not None:
            pv_name = str(name).strip()
            if size is not None:
                try:
                    result[pv_name] = float(size)
                except (ValueError, TypeError):
                    report.warnings.append(
                        f"FBR: non-numeric max line size for '{pv_name}' "
                        f"(source row {r}): {size!r} — skipped."
                    )
    return result


def write_leaks(
    tgt_wb: Workbook,
    pv_records: list[PVLeakRecord],
    opts: LeakOptions,
    fbr_diameters: dict[str, float],
    report: TransferReport,
) -> None:
    """Write one row per PV × leak size to the target Leak sheet."""
    ws = find_sheet(tgt_wb, L.LEAK_SHEET_NAME)
    first_col_idx, last_col_idx = L.leak_scan_col_indices()

    existing_leak_keys: set[tuple[str, str]] = set()

    if opts.transfer_mode is TransferMode.OVERWRITE:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        if existing_last >= L.DATA_START_ROW:
            clear_range(
                ws, first_col_idx, last_col_idx,
                L.DATA_START_ROW, existing_last,
            )
            report.info.append(
                f"Overwrite: cleared rows {L.DATA_START_ROW}–{existing_last} "
                f"on '{L.LEAK_SHEET_NAME}'."
            )
        start_row = L.DATA_START_ROW
    elif opts.transfer_mode is TransferMode.SKIP_EXISTING:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        existing_leak_keys = _read_existing_leak_keys(ws, existing_last)
        start_row = max(existing_last + 1, L.DATA_START_ROW)
        report.info.append(
            f"Skip existing: {len(existing_leak_keys)} (PV, leak) pair(s) already "
            f"in target will be skipped."
        )
    else:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        start_row = max(existing_last + 1, L.DATA_START_ROW)
        report.info.append(
            f"Append: starting at row {start_row} on '{L.LEAK_SHEET_NAME}'."
        )

    cols = {
        "use": col_letter_to_index(L.LEAK_COL_USE),
        "study": col_letter_to_index(L.LEAK_COL_STUDY),
        "folder": col_letter_to_index(L.LEAK_COL_FOLDER),
        "pv_name": col_letter_to_index(L.LEAK_COL_PV_NAME),
        "name": col_letter_to_index(L.LEAK_COL_NAME),
        "orifice_diameter": col_letter_to_index(L.LEAK_COL_ORIFICE_DIAMETER),
        "outdoor_release": col_letter_to_index(L.LEAK_COL_OUTDOOR_RELEASE_DIRECTION),
    }

    active_leaks = [ls for ls in opts.leak_sizes if ls.name.strip()]
    current_row = start_row
    skip_mode = opts.transfer_mode is TransferMode.SKIP_EXISTING
    skipped_count = 0
    filtered_count = 0

    for pv in pv_records:
        fbr_diameter: float | None = fbr_diameters.get(pv.pv_name) if opts.fbr_enabled else None
        for leak in active_leaks:
            key = (pv.pv_name, leak.name.strip())
            if skip_mode and key in existing_leak_keys:
                skipped_count += 1
                continue

            is_fbr_leak = opts.fbr_enabled and leak.name.strip().upper() == "FBR"

            if is_fbr_leak:
                diameter: float | None = fbr_diameter
                if diameter is None:
                    report.warnings.append(
                        f"FBR: no max line size found for PV '{pv.pv_name}' "
                        f"— orifice diameter left empty."
                    )
            else:
                diameter = leak.orifice_diameter
                if (
                    opts.fbr_enabled
                    and fbr_diameter is not None
                    and diameter is not None
                    and diameter > fbr_diameter
                ):
                    filtered_count += 1
                    continue

            ws.cell(row=current_row, column=cols["use"]).value = pv.use
            ws.cell(row=current_row, column=cols["study"]).value = pv.study
            ws.cell(row=current_row, column=cols["folder"]).value = pv.folder
            ws.cell(row=current_row, column=cols["pv_name"]).value = pv.pv_name
            ws.cell(row=current_row, column=cols["name"]).value = leak.name.strip()
            ws.cell(row=current_row, column=cols["orifice_diameter"]).value = diameter
            ws.cell(row=current_row, column=cols["outdoor_release"]).value = (
                L.LEAK_VALUE_OUTDOOR_RELEASE_DIRECTION
            )
            current_row += 1

    if skip_mode and skipped_count:
        report.info.append(f"Skip existing: {skipped_count} leak row(s) skipped.")
    if filtered_count:
        report.info.append(
            f"FBR filter: {filtered_count} leak row(s) omitted "
            f"(orifice diameter exceeds PV max line size)."
        )

    report.rows_written = current_row - start_row
    if report.rows_written > 0:
        report.first_row = start_row
        report.last_row = current_row - 1


def _read_existing_leak_keys(ws, last_row: int) -> set[tuple[str, str]]:
    """Return {(pv_name, leak_name)} for all populated rows in the Leak sheet."""
    pv_idx   = col_letter_to_index(L.LEAK_COL_PV_NAME)
    name_idx = col_letter_to_index(L.LEAK_COL_NAME)
    keys: set[tuple[str, str]] = set()
    for r in range(L.DATA_START_ROW, last_row + 1):
        pv   = ws.cell(row=r, column=pv_idx).value
        name = ws.cell(row=r, column=name_idx).value
        if pv is not None and name is not None:
            keys.add((str(pv).strip(), str(name).strip()))
    return keys
