"""Mixture: read source streams + compositions → write to target ``MIXTURE`` sheet."""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum

from openpyxl.workbook.workbook import Workbook

from . import target_layout as L
from .columns import col_letter_to_index
from .excel_io import find_sheet
from .pressure_vessel import TransferMode, TransferReport
from .sheet_utils import (
    clear_range,
    find_last_populated_row,
    iter_populated_rows,
)
from .smart_match import ComponentMatcher, MatchResult


class CompositionBasis(Enum):
    MOLE = "mole"
    MASS = "mass"


@dataclass
class MixtureSourceConfig:
    """Where to find mixtures in the source workbook."""

    sheet: str
    header_row: int  # row containing component names across the matrix
    start_row: int  # first row of stream data
    stream_col: str  # column with stream identifiers (e.g. "FULL SECTION ID")
    components_first_col: str  # first column of the contiguous component range
    components_last_col: str  # last column of the contiguous component range


@dataclass
class MixtureOptions:
    composition_basis: CompositionBasis = CompositionBasis.MOLE
    smart_match: bool = True
    user_overrides: dict[str, str] | None = None
    transfer_mode: TransferMode = TransferMode.OVERWRITE
    skip_zero: bool = True
    zero_threshold: float = 0.0  # values <= this are treated as zero
    decimal_places: int | None = None  # None = no rounding


@dataclass
class ComponentEntry:
    component_canonical: str  # what gets written to col G
    composition: float
    matched: bool
    method: str  # "exact" | "normalised" | "smart" | "none"
    source_header: str


@dataclass
class MixtureRecord:
    name: str
    components: list[ComponentEntry] = field(default_factory=list)
    source_row: int = 0


# ---------------------------------------------------------------------------


def read_mixtures(
    src_wb: Workbook,
    cfg: MixtureSourceConfig,
    opts: MixtureOptions,
    report: TransferReport,
) -> list[MixtureRecord]:
    """Read mixture rows + their composition vectors from the source."""
    ws = find_sheet(src_wb, cfg.sheet)

    stream_idx = col_letter_to_index(cfg.stream_col)
    comp_first_idx = col_letter_to_index(cfg.components_first_col)
    comp_last_idx = col_letter_to_index(cfg.components_last_col)
    if comp_last_idx < comp_first_idx:
        raise ValueError(
            "Components last column must be >= first column "
            f"({cfg.components_last_col} < {cfg.components_first_col})"
        )

    # Pull header labels (component names) for the component range.
    header_labels: list[str | None] = [
        _stringify(ws.cell(row=cfg.header_row, column=c).value)
        for c in range(comp_first_idx, comp_last_idx + 1)
    ]

    matcher = ComponentMatcher(
        smart_match=opts.smart_match,
        user_overrides=opts.user_overrides,
    )

    records: list[MixtureRecord] = []
    unmatched_summary: dict[str, set[str]] = {}  # source name -> set(streams)
    seen_names: dict[str, int] = {}              # name -> first source row
    duplicate_rows: dict[str, list[int]] = {}    # name -> skipped rows

    for r in iter_populated_rows(ws, stream_idx, cfg.start_row):
        name_raw = ws.cell(row=r, column=stream_idx).value
        name = str(name_raw).strip()
        if not name:
            continue

        if name in seen_names:
            duplicate_rows.setdefault(name, []).append(r)
            continue

        seen_names[name] = r
        record = MixtureRecord(name=name, source_row=r)

        for offset, header in enumerate(header_labels):
            cell_val = ws.cell(row=r, column=comp_first_idx + offset).value
            comp_value = _coerce_float(
                cell_val,
                f"composition[{header!r}]",
                r,
                report,
                allow_none=True,
            )

            if comp_value is None:
                continue
            if not isinstance(comp_value, (int, float)):
                # _coerce_float returned a string for an unparseable cell;
                # skip it for safety.
                continue
            if opts.skip_zero and float(comp_value) <= opts.zero_threshold:
                continue
            if header is None or not header.strip():
                report.warnings.append(
                    f"Source row {r}: composition value found under empty "
                    f"header at column "
                    f"{cfg.components_first_col}+{offset} — skipped."
                )
                continue

            mr: MatchResult = matcher.match(header)
            if not mr.matched:
                unmatched_summary.setdefault(header, set()).add(name)

            record.components.append(
                ComponentEntry(
                    component_canonical=mr.name,
                    composition=float(comp_value),
                    matched=mr.matched,
                    method=mr.method,
                    source_header=header,
                )
            )

        if record.components:
            records.append(record)
        else:
            report.warnings.append(
                f"Source row {r}: stream {name!r} has no non-zero "
                f"compositions — skipped."
            )

    if duplicate_rows:
        for name, skipped in sorted(duplicate_rows.items()):
            first = seen_names[name]
            skipped_str = ", ".join(str(r) for r in skipped)
            report.warnings.append(
                f"Duplicate stream {name!r}: first occurrence (row {first}) "
                f"transferred; skipped row(s): {skipped_str}."
            )

    if unmatched_summary:
        details = ", ".join(
            f"{src!r} (in {len(streams)} stream"
            f"{'s' if len(streams) != 1 else ''})"
            for src, streams in sorted(unmatched_summary.items())
        )
        report.warnings.append(
            f"Unmatched component name(s) written as-is — review in Phast: "
            f"{details}"
        )

    return records


def write_mixtures(
    tgt_wb: Workbook,
    records: list[MixtureRecord],
    opts: MixtureOptions,
    report: TransferReport,
) -> None:
    """Write mixture records to the target ``MIXTURE`` sheet.

    Each record produces ``len(record.components)`` rows; the first row
    carries the mixture-level metadata (Use, name, etc.) and subsequent rows
    carry only the component + composition.
    """
    ws = find_sheet(tgt_wb, L.MIX_SHEET_NAME)
    first_col_idx, last_col_idx = L.mix_scan_col_indices()

    if opts.transfer_mode is TransferMode.OVERWRITE:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        if existing_last >= L.DATA_START_ROW:
            clear_range(
                ws,
                first_col_idx,
                last_col_idx,
                L.DATA_START_ROW,
                existing_last,
            )
            report.info.append(
                f"Overwrite: cleared rows {L.DATA_START_ROW}–{existing_last} "
                f"on '{L.MIX_SHEET_NAME}'."
            )
        start_row = L.DATA_START_ROW
    else:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        start_row = max(existing_last + 1, L.DATA_START_ROW)
        report.info.append(
            f"Append: starting at row {start_row} on '{L.MIX_SHEET_NAME}'."
        )

    use_mole = opts.composition_basis is CompositionBasis.MOLE
    composition_col = L.MIX_COL_MOLE if use_mole else L.MIX_COL_MASS

    cols = {
        "use": col_letter_to_index(L.MIX_COL_USE),
        "pps": col_letter_to_index(L.MIX_COL_PHYSICAL_PROPERTIES_SYSTEM),
        "materials": col_letter_to_index(L.MIX_COL_MATERIALS),
        "name": col_letter_to_index(L.MIX_COL_NAME),
        "component": col_letter_to_index(L.MIX_COL_COMPONENT),
        "composition": col_letter_to_index(composition_col),
        "property_method": col_letter_to_index(L.MIX_COL_PROPERTY_METHOD),
    }

    dp = opts.decimal_places
    r = start_row
    total_rows = 0
    for record in records:
        for i, entry in enumerate(record.components):
            if i == 0:
                ws.cell(row=r, column=cols["use"]).value = L.MIX_VALUE_USE
                ws.cell(row=r, column=cols["pps"]).value = (
                    L.MIX_VALUE_PHYSICAL_PROPERTIES_SYSTEM
                )
                ws.cell(row=r, column=cols["materials"]).value = (
                    L.MIX_VALUE_MATERIALS
                )
                ws.cell(row=r, column=cols["name"]).value = record.name
                ws.cell(row=r, column=cols["property_method"]).value = (
                    L.MIX_VALUE_PROPERTY_METHOD
                )
            ws.cell(row=r, column=cols["component"]).value = (
                entry.component_canonical
            )
            comp = entry.composition if dp is None else round(entry.composition, dp)
            ws.cell(row=r, column=cols["composition"]).value = comp
            r += 1
            total_rows += 1

    report.rows_written = total_rows
    if total_rows:
        report.first_row = start_row
        report.last_row = start_row + total_rows - 1


# ---------------------------------------------------------------------------


def _stringify(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _coerce_float(
    value, label: str, source_row: int, report: TransferReport, allow_none: bool
):
    if value is None:
        return None if allow_none else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None if allow_none else 0.0
    try:
        return float(s)
    except ValueError:
        report.warnings.append(
            f"Source row {source_row}: cannot parse {label}={value!r} as "
            f"number; skipping."
        )
        return value
