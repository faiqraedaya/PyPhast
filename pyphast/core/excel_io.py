"""Excel workbook loading and saving.

Some Phast-generated workbooks use non-standard internal archive paths
(e.g. ``xl/workbook22.xml`` instead of ``xl/workbook.xml``), which causes
openpyxl to fail. This module transparently normalises such archives in
memory before loading.
"""

from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook as _load_workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType


_WORKBOOK_RE = re.compile(r"^xl/workbook\d*\.xml$")
_WORKBOOK_RELS_RE = re.compile(r"^xl/_rels/workbook\d*\.xml\.rels$")


def _normalise_archive(path: Path) -> bytes | None:
    """If the xlsx archive uses non-standard workbook paths, return a
    normalised in-memory copy. Returns None if the archive is already
    standard.
    """
    with zipfile.ZipFile(path, "r") as zin:
        names = zin.namelist()

        wb_name = next((n for n in names if _WORKBOOK_RE.match(n)), None)
        wb_rels_name = next(
            (n for n in names if _WORKBOOK_RELS_RE.match(n)), None
        )

        if wb_name == "xl/workbook.xml" and (
            wb_rels_name is None or wb_rels_name == "xl/_rels/workbook.xml.rels"
        ):
            return None  # already standard

        if wb_name is None:
            raise ValueError(f"No workbook XML found in archive: {path}")

        old_basename = wb_name.split("/")[-1]  # e.g. "workbook22.xml"

        out = io.BytesIO()
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                new_name = item.filename
                if item.filename == wb_name:
                    new_name = "xl/workbook.xml"
                elif (
                    wb_rels_name is not None
                    and item.filename == wb_rels_name
                ):
                    new_name = "xl/_rels/workbook.xml.rels"
                elif item.filename in ("[Content_Types].xml", "_rels/.rels"):
                    data = data.replace(
                        old_basename.encode(), b"workbook.xml"
                    )
                zout.writestr(new_name, data)

        return out.getvalue()


def load_workbook(path: str | Path, data_only: bool = False) -> WorkbookType:
    """Load an xlsx workbook, transparently handling non-standard
    Phast-style archives. Always loads in writeable (non-read-only) mode.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {p}")

    patched = _normalise_archive(p)
    source: Path | io.BytesIO = io.BytesIO(patched) if patched else p
    return _load_workbook(source, data_only=data_only)


def save_workbook(wb: WorkbookType, path: str | Path) -> None:
    """Save workbook to disk. openpyxl writes a standard archive."""
    wb.save(str(path))


def find_sheet(wb: WorkbookType, name: str):
    """Locate a sheet by name, case-insensitive. Raises KeyError if missing."""
    target = name.strip().lower()
    for sn in wb.sheetnames:
        if sn.lower() == target:
            return wb[sn]
    raise KeyError(
        f"Sheet {name!r} not found. Available sheets: {wb.sheetnames}"
    )


def list_sheet_names(path: str | Path) -> list[str]:
    """Return sheet names without keeping the workbook in memory."""
    wb = load_workbook(path)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()
