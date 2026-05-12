"""Pressure vessel: read source rows → write to target ``Pressure vessel`` sheet."""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum

from openpyxl.workbook.workbook import Workbook

from . import target_layout as L
from .columns import col_letter_to_index
from .excel_io import find_sheet
from .sheet_utils import (
    clear_range,
    find_last_populated_row,
    iter_populated_rows,
)


class InventoryMode(Enum):
    MASS = "mass"
    VOLUME = "volume"


class TransferMode(Enum):
    OVERWRITE = "overwrite"
    APPEND = "append"
    SKIP_EXISTING = "skip_existing"


@dataclass
class PressureVesselSourceConfig:
    """Where to find pressure vessel data in the source workbook."""

    sheet: str
    start_row: int
    name_col: str
    stream_col: str
    pressure_col: str
    temperature_col: str
    inventory_col: str  # ignored if assume_unlimited=True


@dataclass
class PressureVesselOptions:
    inventory_mode: InventoryMode = InventoryMode.MASS
    assume_unlimited: bool = False
    transfer_mode: TransferMode = TransferMode.OVERWRITE
    decimal_places: int | None = None  # None = no rounding


@dataclass
class PressureVesselRow:
    name: str
    stream: str | None
    pressure: float | None
    temperature: float | None
    inventory: float | None  # may be None if assume_unlimited
    source_row: int


@dataclass
class TransferReport:
    rows_written: int = 0
    first_row: int | None = None
    last_row: int | None = None
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    info: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------


def read_pressure_vessels(
    src_wb: Workbook,
    cfg: PressureVesselSourceConfig,
    opts: PressureVesselOptions,
    report: TransferReport,
) -> list[PressureVesselRow]:
    """Read pressure vessel rows from the source workbook."""
    ws = find_sheet(src_wb, cfg.sheet)

    name_idx = col_letter_to_index(cfg.name_col)
    stream_idx = col_letter_to_index(cfg.stream_col)
    pressure_idx = col_letter_to_index(cfg.pressure_col)
    temperature_idx = col_letter_to_index(cfg.temperature_col)
    inventory_idx = (
        None if opts.assume_unlimited else col_letter_to_index(cfg.inventory_col)
    )

    rows: list[PressureVesselRow] = []
    for r in iter_populated_rows(ws, name_idx, cfg.start_row):
        name = ws.cell(row=r, column=name_idx).value
        section_name = str(name).strip() if name is not None else ""
        stream = ws.cell(row=r, column=stream_idx).value
        pressure = _coerce_float(
            ws.cell(row=r, column=pressure_idx).value, "pressure", r, report,
            section_name=section_name,
        )
        temperature = _coerce_float(
            ws.cell(row=r, column=temperature_idx).value,
            "temperature",
            r,
            report,
            section_name=section_name,
        )
        inventory: float | None
        if opts.assume_unlimited:
            inventory = None
        else:
            assert inventory_idx is not None  # for the type checker
            inventory = _coerce_float(
                ws.cell(row=r, column=inventory_idx).value,
                "inventory",
                r,
                report,
                section_name=section_name,
            )

        rows.append(
            PressureVesselRow(
                name=str(name).strip(),
                stream=None if stream is None else str(stream).strip(),
                pressure=pressure,
                temperature=temperature,
                inventory=inventory,
                source_row=r,
            )
        )
    return rows


def write_pressure_vessels(
    tgt_wb: Workbook,
    rows: list[PressureVesselRow],
    opts: PressureVesselOptions,
    report: TransferReport,
) -> None:
    """Write rows to the target ``Pressure vessel`` sheet."""
    ws = find_sheet(tgt_wb, L.PV_SHEET_NAME)
    first_col_idx, last_col_idx = L.pv_scan_col_indices()

    if opts.transfer_mode is TransferMode.OVERWRITE:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        if existing_last >= L.DATA_START_ROW:
            clear_range(
                ws,
                first_col_idx,
                last_col_idx,
                L.DATA_START_ROW,
                existing_last,
            )
            report.info.append(
                f"Overwrite: cleared rows {L.DATA_START_ROW}–{existing_last} "
                f"on '{L.PV_SHEET_NAME}'."
            )
        start_row = L.DATA_START_ROW
    elif opts.transfer_mode is TransferMode.SKIP_EXISTING:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        existing_names = _read_existing_names(ws, col_letter_to_index(L.PV_COL_NAME),
                                              first_col_idx, last_col_idx)
        original_count = len(rows)
        rows = [r for r in rows if r.name not in existing_names]
        skipped = original_count - len(rows)
        start_row = max(existing_last + 1, L.DATA_START_ROW)
        report.info.append(
            f"Skip existing: {skipped} PV(s) already in target skipped, "
            f"{len(rows)} new PV(s) to append."
        )
    else:  # APPEND
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        start_row = max(existing_last + 1, L.DATA_START_ROW)
        report.info.append(
            f"Append: starting at row {start_row} on '{L.PV_SHEET_NAME}'."
        )

    inv_value_unlimited = L.UNLIMITED_INVENTORY_VALUE
    use_volume = opts.inventory_mode is InventoryMode.VOLUME
    specify_volume_value = (
        L.PV_VALUE_VOLUME_YES if use_volume else L.PV_VALUE_VOLUME_NO
    )
    inv_target_col = (
        L.PV_COL_VOLUME_INVENTORY if use_volume else L.PV_COL_MASS_INVENTORY
    )

    cols = {
        "use": col_letter_to_index(L.PV_COL_USE),
        "name": col_letter_to_index(L.PV_COL_NAME),
        "material": col_letter_to_index(L.PV_COL_MATERIAL),
        "specify_volume": col_letter_to_index(L.PV_COL_SPECIFY_VOLUME),
        "inventory": col_letter_to_index(inv_target_col),
        "material_to_track": col_letter_to_index(L.PV_COL_MATERIAL_TO_TRACK),
        "temperature": col_letter_to_index(L.PV_COL_TEMPERATURE),
        "pressure": col_letter_to_index(L.PV_COL_PRESSURE_GAUGE),
    }

    dp = opts.decimal_places
    for offset, row in enumerate(rows):
        r = start_row + offset
        ws.cell(row=r, column=cols["use"]).value = L.PV_VALUE_USE
        ws.cell(row=r, column=cols["name"]).value = row.name
        ws.cell(row=r, column=cols["material"]).value = row.stream
        ws.cell(row=r, column=cols["material_to_track"]).value = row.stream
        ws.cell(row=r, column=cols["specify_volume"]).value = (
            specify_volume_value
        )
        ws.cell(row=r, column=cols["inventory"]).value = (
            inv_value_unlimited if opts.assume_unlimited
            else _maybe_round(row.inventory, dp)
        )
        ws.cell(row=r, column=cols["temperature"]).value = _maybe_round(row.temperature, dp)
        ws.cell(row=r, column=cols["pressure"]).value = _maybe_round(row.pressure, dp)

    report.rows_written = len(rows)
    if rows:
        report.first_row = start_row
        report.last_row = start_row + len(rows) - 1


# ---------------------------------------------------------------------------


def _read_existing_names(ws, name_col_idx: int, first_c: int, last_c: int) -> set[str]:
    """Read non-empty name values from the target sheet into a set."""
    last_row = find_last_populated_row(ws, first_c, last_c, L.DATA_START_ROW)
    existing: set[str] = set()
    for r in range(L.DATA_START_ROW, last_row + 1):
        v = ws.cell(row=r, column=name_col_idx).value
        if v is not None:
            s = str(v).strip()
            if s:
                existing.add(s)
    return existing


def _maybe_round(value: float | None, places: int | None) -> float | None:
    if places is None or value is None:
        return value
    return round(float(value), places)


def _coerce_float(
    value, label: str, source_row: int, report: TransferReport,
    section_name: str = ""
) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        name_part = f" '{section_name}'" if section_name else ""
        report.warnings.append(
            f"Section{name_part} (row {source_row}): {label} value {value!r} "
            f"is not numeric — defaulted to 0."
        )
        return 0.0
