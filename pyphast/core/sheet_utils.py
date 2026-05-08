"""Helpers for scanning, clearing and iterating rectangular ranges."""

from __future__ import annotations

from typing import Iterable

from openpyxl.worksheet.worksheet import Worksheet


def _is_populated(v) -> bool:
    if v is None:
        return False
    if isinstance(v, str) and v.strip() == "":
        return False
    return True


def find_last_populated_row(
    ws: Worksheet,
    first_col: int,
    last_col: int,
    start_row: int,
    hard_stop: int = 100_000,
) -> int:
    """Return the highest 1-based row index ``r >= start_row`` such that any
    cell in ``ws[first_col..last_col, r]`` is non-empty.

    Returns ``start_row - 1`` if nothing is populated — so ``last + 1`` is
    always the correct next-free row.
    """
    last = start_row - 1
    upper = min(ws.max_row or start_row, hard_stop)
    if upper < start_row:
        return last
    for r in range(start_row, upper + 1):
        for c in range(first_col, last_col + 1):
            if _is_populated(ws.cell(row=r, column=c).value):
                last = r
                break
    return last


def clear_range(
    ws: Worksheet,
    first_col: int,
    last_col: int,
    first_row: int,
    last_row: int,
) -> None:
    """Set values to None across the rectangle. Formatting is preserved."""
    if last_row < first_row:
        return
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            ws.cell(row=r, column=c).value = None


def read_column_values(
    ws: Worksheet,
    col_index: int,
    start_row: int,
    end_row: int | None = None,
) -> list:
    """Read column values from start_row to end_row (inclusive)."""
    if end_row is None:
        end_row = ws.max_row or start_row
    return [
        ws.cell(row=r, column=col_index).value
        for r in range(start_row, end_row + 1)
    ]


def iter_populated_rows(
    ws: Worksheet,
    key_col: int,
    start_row: int,
    end_row: int | None = None,
) -> Iterable[int]:
    """Yield row indices where the cell in ``key_col`` is non-empty."""
    if end_row is None:
        end_row = ws.max_row or start_row
    for r in range(start_row, end_row + 1):
        if _is_populated(ws.cell(row=r, column=key_col).value):
            yield r
