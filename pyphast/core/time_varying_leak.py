"""Time varying leak: read PV names from target PV sheet → write to target TVL sheet."""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.workbook.workbook import Workbook

from . import target_layout as L
from .columns import col_letter_to_index
from .excel_io import find_sheet
from .leak import LeakSizeEntry, LeakSourceConfig, PVLeakRecord, read_fbr_diameters
from .pressure_vessel import TransferMode, TransferReport
from .sheet_utils import clear_range, find_last_populated_row, iter_populated_rows

_VAPOUR_PHASES = {"V", "SC"}


@dataclass
class TVLOptions:
    leak_sizes: list[LeakSizeEntry]
    fbr_enabled: bool = False
    transfer_mode: TransferMode = TransferMode.OVERWRITE
    avg_rate_method: str = L.TVL_AVG_RATE_METHOD_DEFAULT
    safety_system: str = L.TVL_SAFETY_SYSTEM_YES
    isolation: str = L.TVL_ISOLATION_SUCCESSFUL
    time_to_isolation: float | None = None


# ---------------------------------------------------------------------------


def read_pv_phases_from_source(
    src_wb: Workbook,
    sheet: str,
    name_col: str,
    phase_col: str,
    start_row: int,
) -> dict[str, str]:
    """Return {pv_name: phase_string} from the source workbook."""
    ws = find_sheet(src_wb, sheet)
    name_idx = col_letter_to_index(name_col)
    phase_idx = col_letter_to_index(phase_col)
    result: dict[str, str] = {}
    for r in iter_populated_rows(ws, name_idx, start_row):
        name = ws.cell(row=r, column=name_idx).value
        phase = ws.cell(row=r, column=phase_idx).value
        if name is not None:
            pv_name = str(name).strip()
            result[pv_name] = str(phase).strip().upper() if phase is not None else ""
    return result


def is_vapour_phase(phase: str) -> bool:
    return phase.upper() in _VAPOUR_PHASES


def split_pv_records_by_phase(
    pv_records: list[PVLeakRecord],
    phases: dict[str, str],
    report: TransferReport,
) -> tuple[list[PVLeakRecord], list[PVLeakRecord]]:
    """Split records into (liquid_records, vapour_records) based on phase lookup.

    PVs whose phase is V or SC → vapour_records (TVL).
    All others → liquid_records (Leak).
    PVs with no phase entry → liquid_records with a warning.
    """
    liquid: list[PVLeakRecord] = []
    vapour: list[PVLeakRecord] = []
    for pv in pv_records:
        phase = phases.get(pv.pv_name, "")
        if not phase:
            report.warnings.append(
                f"Phase not found for PV '{pv.pv_name}' — defaulting to Leak (liquid)."
            )
            liquid.append(pv)
        elif is_vapour_phase(phase):
            vapour.append(pv)
        else:
            liquid.append(pv)
    return liquid, vapour


# ---------------------------------------------------------------------------


def write_tvls(
    tgt_wb: Workbook,
    pv_records: list[PVLeakRecord],
    opts: TVLOptions,
    fbr_diameters: dict[str, float],
    report: TransferReport,
) -> None:
    """Write one row per PV × leak size to the target Time varying leak sheet."""
    ws = find_sheet(tgt_wb, L.TVL_SHEET_NAME)
    first_col_idx, last_col_idx = L.tvl_scan_col_indices()

    existing_tvl_keys: set[tuple[str, str]] = set()

    if opts.transfer_mode is TransferMode.OVERWRITE:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        if existing_last >= L.DATA_START_ROW:
            clear_range(
                ws, first_col_idx, last_col_idx,
                L.DATA_START_ROW, existing_last,
            )
            report.info.append(
                f"Overwrite: cleared rows {L.DATA_START_ROW}–{existing_last} "
                f"on '{L.TVL_SHEET_NAME}'."
            )
        start_row = L.DATA_START_ROW
    elif opts.transfer_mode is TransferMode.SKIP_EXISTING:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        existing_tvl_keys = _read_existing_tvl_keys(ws, existing_last)
        start_row = max(existing_last + 1, L.DATA_START_ROW)
        report.info.append(
            f"Skip existing: {len(existing_tvl_keys)} (PV, TVL) pair(s) already "
            f"in target will be skipped."
        )
    else:
        existing_last = find_last_populated_row(
            ws, first_col_idx, last_col_idx, L.DATA_START_ROW
        )
        start_row = max(existing_last + 1, L.DATA_START_ROW)
        report.info.append(
            f"Append: starting at row {start_row} on '{L.TVL_SHEET_NAME}'."
        )

    cols = {
        "use": col_letter_to_index(L.TVL_COL_USE),
        "study": col_letter_to_index(L.TVL_COL_STUDY),
        "folder": col_letter_to_index(L.TVL_COL_FOLDER),
        "pv_name": col_letter_to_index(L.TVL_COL_PV_NAME),
        "name": col_letter_to_index(L.TVL_COL_NAME),
        "orifice_diameter": col_letter_to_index(L.TVL_COL_ORIFICE_DIAMETER),
        "outdoor_release": col_letter_to_index(L.TVL_COL_OUTDOOR_RELEASE_DIRECTION),
        "avg_rate_method": col_letter_to_index(L.TVL_COL_AVG_RATE_METHOD),
        "safety_system": col_letter_to_index(L.TVL_COL_SAFETY_SYSTEM),
        "isolation": col_letter_to_index(L.TVL_COL_ISOLATION),
        "time_to_isolation": col_letter_to_index(L.TVL_COL_TIME_TO_ISOLATION),
    }

    safety_yes = opts.safety_system == L.TVL_SAFETY_SYSTEM_YES
    isolation_success = opts.isolation == L.TVL_ISOLATION_SUCCESSFUL

    active_leaks = [ls for ls in opts.leak_sizes if ls.name.strip()]
    current_row = start_row
    skip_mode = opts.transfer_mode is TransferMode.SKIP_EXISTING
    skipped_count = 0
    filtered_count = 0

    for pv in pv_records:
        fbr_diameter: float | None = fbr_diameters.get(pv.pv_name) if opts.fbr_enabled else None
        for leak in active_leaks:
            key = (pv.pv_name, leak.name.strip())
            if skip_mode and key in existing_tvl_keys:
                skipped_count += 1
                continue

            is_fbr_leak = opts.fbr_enabled and leak.name.strip().upper() == "FBR"

            if is_fbr_leak:
                diameter: float | None = fbr_diameter
                if diameter is None:
                    report.warnings.append(
                        f"FBR: no max line size found for PV '{pv.pv_name}' "
                        f"— orifice diameter left empty."
                    )
            else:
                diameter = leak.orifice_diameter
                if (
                    opts.fbr_enabled
                    and fbr_diameter is not None
                    and diameter is not None
                    and diameter > fbr_diameter
                ):
                    filtered_count += 1
                    continue

            ws.cell(row=current_row, column=cols["use"]).value = pv.use
            ws.cell(row=current_row, column=cols["study"]).value = pv.study
            ws.cell(row=current_row, column=cols["folder"]).value = pv.folder
            ws.cell(row=current_row, column=cols["pv_name"]).value = pv.pv_name
            ws.cell(row=current_row, column=cols["name"]).value = leak.name.strip()
            ws.cell(row=current_row, column=cols["orifice_diameter"]).value = diameter
            ws.cell(row=current_row, column=cols["outdoor_release"]).value = (
                L.TVL_VALUE_OUTDOOR_RELEASE_DIRECTION
            )
            ws.cell(row=current_row, column=cols["avg_rate_method"]).value = (
                opts.avg_rate_method
            )
            ws.cell(row=current_row, column=cols["safety_system"]).value = (
                opts.safety_system
            )
            if safety_yes:
                ws.cell(row=current_row, column=cols["isolation"]).value = (
                    opts.isolation
                )
                if isolation_success and opts.time_to_isolation is not None:
                    ws.cell(row=current_row, column=cols["time_to_isolation"]).value = (
                        opts.time_to_isolation
                    )
            current_row += 1

    if skip_mode and skipped_count:
        report.info.append(f"Skip existing: {skipped_count} TVL row(s) skipped.")
    if filtered_count:
        report.info.append(
            f"FBR filter: {filtered_count} TVL row(s) omitted "
            f"(orifice diameter exceeds PV max line size)."
        )

    report.rows_written = current_row - start_row
    if report.rows_written > 0:
        report.first_row = start_row
        report.last_row = current_row - 1


def _read_existing_tvl_keys(ws, last_row: int) -> set[tuple[str, str]]:
    """Return {(pv_name, tvl_name)} for all populated rows in the TVL sheet."""
    pv_idx   = col_letter_to_index(L.TVL_COL_PV_NAME)
    name_idx = col_letter_to_index(L.TVL_COL_NAME)
    keys: set[tuple[str, str]] = set()
    for r in range(L.DATA_START_ROW, last_row + 1):
        pv   = ws.cell(row=r, column=pv_idx).value
        name = ws.cell(row=r, column=name_idx).value
        if pv is not None and name is not None:
            keys.add((str(pv).strip(), str(name).strip()))
    return keys
