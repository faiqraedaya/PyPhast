"""Tests for pyphast.core.sheet_utils."""

import pytest
from openpyxl import Workbook

from pyphast.core.sheet_utils import (
    clear_range,
    find_last_populated_row,
    iter_populated_rows,
    read_column_values,
)


def _ws(data: dict[tuple[int, int], object]):
    """Return an in-memory worksheet pre-populated with {(row, col): value}."""
    wb = Workbook()
    ws = wb.active
    for (r, c), v in data.items():
        ws.cell(row=r, column=c).value = v
    return ws


class TestFindLastPopulatedRow:
    def test_empty_sheet_returns_before_start(self):
        ws = _ws({})
        assert find_last_populated_row(ws, 1, 3, start_row=5) == 4

    def test_single_row(self):
        ws = _ws({(5, 1): "hello"})
        assert find_last_populated_row(ws, 1, 1, start_row=5) == 5

    def test_multiple_rows_returns_last(self):
        ws = _ws({(5, 1): "a", (6, 1): "b", (7, 1): "c"})
        assert find_last_populated_row(ws, 1, 1, start_row=5) == 7

    def test_gap_in_middle_still_returns_last(self):
        ws = _ws({(5, 1): "a", (7, 1): "c"})  # row 6 empty
        assert find_last_populated_row(ws, 1, 1, start_row=5) == 7

    def test_rows_before_start_ignored(self):
        ws = _ws({(3, 1): "before", (8, 1): "after"})
        assert find_last_populated_row(ws, 1, 1, start_row=5) == 8

    def test_multi_column_scan(self):
        # Value only in col 3, not col 1 — should still be found
        ws = _ws({(6, 3): "data"})
        assert find_last_populated_row(ws, 1, 3, start_row=5) == 6

    def test_whitespace_value_not_counted(self):
        ws = _ws({(5, 1): "   "})
        # Whitespace-only string is not populated
        assert find_last_populated_row(ws, 1, 1, start_row=5) == 4

    def test_none_value_not_counted(self):
        ws = _ws({(5, 1): None})
        assert find_last_populated_row(ws, 1, 1, start_row=5) == 4

    def test_zero_is_populated(self):
        ws = _ws({(5, 1): 0})
        assert find_last_populated_row(ws, 1, 1, start_row=5) == 5

    def test_false_is_populated(self):
        ws = _ws({(5, 1): False})
        assert find_last_populated_row(ws, 1, 1, start_row=5) == 5


class TestClearRange:
    def test_clears_cells_in_range(self):
        ws = _ws({(2, 1): "a", (2, 2): "b", (3, 1): "c"})
        clear_range(ws, first_col=1, last_col=2, first_row=2, last_row=3)
        assert ws.cell(2, 1).value is None
        assert ws.cell(2, 2).value is None
        assert ws.cell(3, 1).value is None

    def test_does_not_touch_outside_range(self):
        ws = _ws({(1, 1): "keep", (2, 1): "clear", (3, 1): "keep"})
        clear_range(ws, first_col=1, last_col=1, first_row=2, last_row=2)
        assert ws.cell(1, 1).value == "keep"
        assert ws.cell(2, 1).value is None
        assert ws.cell(3, 1).value == "keep"

    def test_noop_when_last_before_first(self):
        ws = _ws({(2, 1): "intact"})
        clear_range(ws, first_col=1, last_col=1, first_row=5, last_row=4)
        assert ws.cell(2, 1).value == "intact"


class TestReadColumnValues:
    def test_basic_read(self):
        ws = _ws({(1, 2): "x", (2, 2): "y", (3, 2): "z"})
        assert read_column_values(ws, col_index=2, start_row=1, end_row=3) == [
            "x",
            "y",
            "z",
        ]

    def test_includes_none_cells(self):
        ws = _ws({(1, 1): "a", (3, 1): "c"})
        result = read_column_values(ws, col_index=1, start_row=1, end_row=3)
        assert result == ["a", None, "c"]

    def test_single_row(self):
        ws = _ws({(5, 3): 42})
        assert read_column_values(ws, col_index=3, start_row=5, end_row=5) == [42]

    def test_end_row_defaults_to_max_row(self):
        ws = _ws({(1, 1): "a", (2, 1): "b"})
        result = read_column_values(ws, col_index=1, start_row=1)
        assert "a" in result
        assert "b" in result


class TestIterPopulatedRows:
    def test_yields_populated_rows(self):
        ws = _ws({(2, 1): "a", (3, 1): None, (4, 1): "b"})
        assert list(iter_populated_rows(ws, key_col=1, start_row=2)) == [2, 4]

    def test_empty_sheet_yields_nothing(self):
        ws = _ws({})
        assert list(iter_populated_rows(ws, key_col=1, start_row=1)) == []

    def test_skips_whitespace_only(self):
        ws = _ws({(1, 1): "  ", (2, 1): "data"})
        assert list(iter_populated_rows(ws, key_col=1, start_row=1)) == [2]

    def test_respects_start_row(self):
        ws = _ws({(1, 1): "before", (3, 1): "after"})
        assert list(iter_populated_rows(ws, key_col=1, start_row=2)) == [3]

    def test_respects_end_row(self):
        ws = _ws({(1, 1): "a", (2, 1): "b", (3, 1): "c"})
        assert list(iter_populated_rows(ws, key_col=1, start_row=1, end_row=2)) == [
            1,
            2,
        ]

    def test_zero_value_counts_as_populated(self):
        ws = _ws({(1, 1): 0})
        assert list(iter_populated_rows(ws, key_col=1, start_row=1)) == [1]
