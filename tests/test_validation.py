"""Tests for pyphast.core.validation."""

import pytest
from openpyxl import Workbook

from pyphast.core.mixture import MixtureSourceConfig
from pyphast.core.pressure_vessel import PressureVesselSourceConfig, TransferReport
from pyphast.core.validation import (
    collect_mixture_stream_names,
    collect_pv_stream_refs,
    warn_unresolved_pv_streams,
)


def _mix_wb(streams: list[str], sheet: str = "Mixtures", stream_col: str = "A", start_row: int = 2) -> Workbook:
    """Build an in-memory workbook with mixture stream names in one column."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for i, name in enumerate(streams):
        ws.cell(row=start_row + i, column=1).value = name
    return wb


def _pv_wb(
    rows: list[tuple[str, str]],  # (pv_name, stream_ref)
    sheet: str = "PVData",
    name_col: str = "A",
    stream_col: str = "B",
    start_row: int = 2,
) -> Workbook:
    """Build an in-memory workbook with PV name and stream reference columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for i, (name, stream) in enumerate(rows):
        ws.cell(row=start_row + i, column=1).value = name
        ws.cell(row=start_row + i, column=2).value = stream
    return wb


class TestCollectMixtureStreamNames:
    def _cfg(self, sheet="Mixtures"):
        return MixtureSourceConfig(
            sheet=sheet,
            header_row=1,
            start_row=2,
            stream_col="A",
            components_first_col="B",
            components_last_col="D",
        )

    def test_returns_stream_names(self):
        wb = _mix_wb(["FEED", "PRODUCT", "RECYCLE"])
        result = collect_mixture_stream_names(wb, self._cfg())
        assert result == {"FEED", "PRODUCT", "RECYCLE"}

    def test_empty_sheet_returns_empty_set(self):
        wb = _mix_wb([])
        result = collect_mixture_stream_names(wb, self._cfg())
        assert result == set()

    def test_ignores_none_cells(self):
        wb = _mix_wb(["FEED", None, "RECYCLE"])
        result = collect_mixture_stream_names(wb, self._cfg())
        assert result == {"FEED", "RECYCLE"}

    def test_ignores_whitespace_only_cells(self):
        wb = _mix_wb(["FEED", "   ", "RECYCLE"])
        result = collect_mixture_stream_names(wb, self._cfg())
        assert result == {"FEED", "RECYCLE"}

    def test_deduplicates_names(self):
        wb = _mix_wb(["FEED", "FEED", "PRODUCT"])
        result = collect_mixture_stream_names(wb, self._cfg())
        assert result == {"FEED", "PRODUCT"}

    def test_case_insensitive_lookup_sheet_name(self):
        wb = _mix_wb(["FEED"], sheet="MIXTURES")
        cfg = MixtureSourceConfig(
            sheet="mixtures",  # lowercase — find_sheet is case-insensitive
            header_row=1,
            start_row=2,
            stream_col="A",
            components_first_col="B",
            components_last_col="D",
        )
        result = collect_mixture_stream_names(wb, cfg)
        assert "FEED" in result

    def test_sheet_not_found_raises(self):
        wb = _mix_wb(["FEED"])
        cfg = MixtureSourceConfig(
            sheet="DoesNotExist",
            header_row=1,
            start_row=2,
            stream_col="A",
            components_first_col="B",
            components_last_col="D",
        )
        with pytest.raises(KeyError):
            collect_mixture_stream_names(wb, cfg)


class TestCollectPvStreamRefs:
    def _cfg(self, sheet="PVData"):
        return PressureVesselSourceConfig(
            sheet=sheet,
            start_row=2,
            name_col="A",
            stream_col="B",
            pressure_col="C",
            temperature_col="D",
            inventory_col="E",
        )

    def test_returns_stream_refs(self):
        wb = _pv_wb([("PV-01", "FEED"), ("PV-02", "PRODUCT")])
        result = collect_pv_stream_refs(wb, self._cfg())
        assert result == {"FEED", "PRODUCT"}

    def test_empty_returns_empty_set(self):
        wb = _pv_wb([])
        result = collect_pv_stream_refs(wb, self._cfg())
        assert result == set()

    def test_ignores_rows_where_name_is_empty(self):
        # name_col empty → row not populated → stream ref skipped
        wb = _pv_wb([(None, "FEED"), ("PV-02", "PRODUCT")])
        result = collect_pv_stream_refs(wb, self._cfg())
        assert result == {"PRODUCT"}

    def test_ignores_none_stream_ref(self):
        wb = _pv_wb([("PV-01", None), ("PV-02", "PRODUCT")])
        result = collect_pv_stream_refs(wb, self._cfg())
        assert result == {"PRODUCT"}

    def test_deduplicates_refs(self):
        wb = _pv_wb([("PV-01", "FEED"), ("PV-02", "FEED")])
        result = collect_pv_stream_refs(wb, self._cfg())
        assert result == {"FEED"}


class TestWarnUnresolvedPvStreams:
    def test_no_warning_when_all_resolved(self):
        report = TransferReport()
        warn_unresolved_pv_streams({"FEED", "PRODUCT"}, {"FEED", "PRODUCT"}, report)
        assert report.warnings == []

    def test_no_warning_when_pv_refs_empty(self):
        report = TransferReport()
        warn_unresolved_pv_streams(set(), {"FEED"}, report)
        assert report.warnings == []

    def test_warning_when_pv_ref_missing_from_mixtures(self):
        report = TransferReport()
        warn_unresolved_pv_streams({"FEED", "MYSTERY"}, {"FEED"}, report)
        assert len(report.warnings) == 1
        assert "MYSTERY" in report.warnings[0]

    def test_warning_lists_all_missing_streams(self):
        report = TransferReport()
        warn_unresolved_pv_streams({"A", "B", "C"}, set(), report)
        assert len(report.warnings) == 1
        assert "A" in report.warnings[0]
        assert "B" in report.warnings[0]
        assert "C" in report.warnings[0]

    def test_warning_message_suggests_mixtures_tab(self):
        report = TransferReport()
        warn_unresolved_pv_streams({"MISSING"}, set(), report)
        assert "Mixtures" in report.warnings[0] or "mixtures" in report.warnings[0].lower()

    def test_extra_mixtures_cause_no_warning(self):
        report = TransferReport()
        warn_unresolved_pv_streams({"FEED"}, {"FEED", "EXTRA1", "EXTRA2"}, report)
        assert report.warnings == []
